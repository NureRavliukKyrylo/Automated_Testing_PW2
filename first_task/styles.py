from openpyxl.styles import PatternFill, Font


class Styles:
    NUMBER = PatternFill(start_color="008000",
                        end_color="008000",
                        fill_type="solid") 
    TEXT = PatternFill(start_color="0000FF",
                        end_color="0000FF",
                        fill_type="solid")   
    MIXED = PatternFill(start_color="FFFF00",
                        end_color="FFFF00",
                        fill_type="solid")