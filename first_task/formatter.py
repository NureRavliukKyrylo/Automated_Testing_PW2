from openpyxl import load_workbook
from styles import Styles
from factory import RuleFactory


class ExcelFormatter:

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.rules = RuleFactory.create_rules()

        self.style_map = {
            "NUMBER": lambda cell: setattr(cell, "fill", Styles.NUMBER),
            "TEXT": lambda cell: setattr(cell, "fill", Styles.TEXT),
            "MIXED": lambda cell: setattr(cell, "fill", Styles.MIXED)
        }

    def format(self):

        workbook = load_workbook(self.file_path)
        sheet = workbook.active

        for row in sheet.iter_rows():
            for cell in row:

                if cell.value is None:
                    continue

                value = str(cell.value)

                for rule, style_key in self.rules.items():
                    if rule.matches(value):
                        self.style_map[style_key](cell)
                        break

        workbook.save(self.file_path)
        workbook.close()